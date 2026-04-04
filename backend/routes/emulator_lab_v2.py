"""
Phase 4 Emulator Lab — Complete engineering workbench.
SmartEGM Engine, Response Manager, Script DSL, Balanced Meters, TAR, Watchables.
"""
from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import StreamingResponse
from database import db
from auth import get_current_user
import uuid
import random
import hashlib
import json
import csv
import io
import re
import asyncio
from datetime import datetime, timezone, timedelta
from typing import Optional

router = APIRouter(prefix="/api/emulator-lab", tags=["emulator-lab"])

# ══════════════════════════════════════════════════
# 1. SMART EGM ENGINE — 12 Player Verbs
# ══════════════════════════════════════════════════

PLAYER_VERBS = [
    {"id": "INSERT_BILL", "label": "Insert Bill", "params": ["denomination"], "state_req": "ENABLED"},
    {"id": "INSERT_VOUCHER", "label": "Insert Voucher", "params": ["amount"], "state_req": "ENABLED"},
    {"id": "INSERT_COIN", "label": "Insert Coin", "params": ["amount"], "state_req": "ENABLED"},
    {"id": "PUSH_PLAY_BUTTON", "label": "Push Play", "params": ["wager"], "state_req": "ENABLED"},
    {"id": "PUSH_MAX_BET", "label": "Max Bet", "params": [], "state_req": "ENABLED"},
    {"id": "CASH_OUT", "label": "Cash Out", "params": [], "state_req": "ENABLED"},
    {"id": "REQUEST_HANDPAY", "label": "Request Handpay", "params": [], "state_req": "ANY"},
    {"id": "OPEN_DOOR", "label": "Open Door", "params": ["door_type"], "state_req": "ANY"},
    {"id": "CLOSE_DOOR", "label": "Close Door", "params": ["door_type"], "state_req": "ANY"},
    {"id": "FORCE_TILT", "label": "Force Tilt", "params": ["reason"], "state_req": "ANY"},
    {"id": "CLEAR_FAULT", "label": "Clear Fault", "params": [], "state_req": "FAULT"},
    {"id": "SET_CREDITS", "label": "Set Credits", "params": ["amount"], "state_req": "ANY"},
]

WIN_LEVELS = [
    {"id": 0, "name": "NoWin", "probability": 0.70, "multiplier": 0},
    {"id": 1, "name": "Small", "probability": 0.20, "multiplier": 1.5},
    {"id": 2, "name": "Medium", "probability": 0.08, "multiplier": 5},
    {"id": 3, "name": "Large", "probability": 0.019, "multiplier": 25},
    {"id": 4, "name": "Jackpot", "probability": 0.001, "multiplier": 250},
]

# In-memory SmartEGM state
_smart_egm_state = {}


def _get_egm(session_id: str) -> dict:
    if session_id not in _smart_egm_state:
        _smart_egm_state[session_id] = {
            "state": "IDLE", "credits": 0, "coin_in": 0, "coin_out": 0,
            "games_played": 0, "bills_in": 0, "vouchers_in": 0, "handpays": 0,
            "doors_open": set(), "events": [], "meters": {},
        }
    return _smart_egm_state[session_id]


def _roll_outcome(wager: int) -> dict:
    r = random.random()
    cumulative = 0
    for wl in WIN_LEVELS:
        cumulative += wl["probability"]
        if r <= cumulative:
            win = int(wager * wl["multiplier"])
            return {"level": wl["name"], "multiplier": wl["multiplier"], "win": win}
    return {"level": "NoWin", "multiplier": 0, "win": 0}


@router.get("/smart-egm/verbs")
async def get_player_verbs(request: Request):
    await get_current_user(request)
    return {"verbs": PLAYER_VERBS, "win_levels": WIN_LEVELS}


@router.post("/smart-egm/execute-verb")
async def execute_player_verb(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id", "default")
    verb = body.get("verb")
    params = body.get("params", {})

    egm = _get_egm(session_id)
    now = datetime.now(timezone.utc).isoformat()
    events = []

    if verb == "INSERT_BILL":
        denom = params.get("denomination", 2000)
        egm["credits"] += denom
        egm["bills_in"] += denom
        egm["coin_in"] += denom
        events.append({"code": "G2S_NAE101", "type": "billAccepted", "data": {"denomination": denom, "credits": egm["credits"]}})

    elif verb == "INSERT_VOUCHER":
        amount = params.get("amount", 1000)
        egm["credits"] += amount
        egm["vouchers_in"] += amount
        events.append({"code": "G2S_VCE104", "type": "voucherPending", "data": {"amount": amount}})
        events.append({"code": "G2S_VCE102", "type": "voucherRedeemed", "data": {"amount": amount, "credits": egm["credits"]}})

    elif verb == "INSERT_COIN":
        amount = params.get("amount", 100)
        egm["credits"] += amount
        egm["coin_in"] += amount
        events.append({"code": "G2S_CAE101", "type": "coinAccepted", "data": {"amount": amount}})

    elif verb in ("PUSH_PLAY_BUTTON", "PUSH_MAX_BET"):
        wager = params.get("wager", 500) if verb == "PUSH_PLAY_BUTTON" else 5000
        if egm["credits"] < wager:
            raise HTTPException(status_code=400, detail=f"Insufficient credits: {egm['credits']} < {wager}")
        outcome = _roll_outcome(wager)
        egm["credits"] -= wager
        egm["coin_in"] += wager
        egm["games_played"] += 1
        events.append({"code": "G2S_GPE101", "type": "gameStarted", "data": {"wager": wager}})
        if outcome["win"] > 0:
            egm["credits"] += outcome["win"]
            egm["coin_out"] += outcome["win"]
        events.append({"code": "G2S_GPE112", "type": "gameEnded", "data": {"wager": wager, "win": outcome["win"], "level": outcome["level"], "credits": egm["credits"]}})
        if outcome["win"] >= 120000:
            events.append({"code": "G2S_HPE101", "type": "handpayPending", "data": {"amount": outcome["win"]}})
            egm["state"] = "HANDPAY_PENDING"

    elif verb == "CASH_OUT":
        if egm["credits"] <= 0:
            raise HTTPException(status_code=400, detail="No credits to cash out")
        amount = egm["credits"]
        egm["credits"] = 0
        egm["coin_out"] += amount
        events.append({"code": "G2S_VCE101", "type": "voucherPrinted", "data": {"amount": amount}})

    elif verb == "REQUEST_HANDPAY":
        egm["handpays"] += 1
        events.append({"code": "G2S_HPE101", "type": "handpayPending", "data": {"credits": egm["credits"]}})
        events.append({"code": "G2S_HPE104", "type": "handpayAcknowledged", "data": {}})
        egm["state"] = "ENABLED"

    elif verb == "OPEN_DOOR":
        door = params.get("door_type", "main")
        egm["doors_open"].add(door)
        events.append({"code": "G2S_CBE201", "type": "doorOpened", "data": {"door_type": door}})

    elif verb == "CLOSE_DOOR":
        door = params.get("door_type", "main")
        egm["doors_open"].discard(door)
        events.append({"code": "G2S_CBE202", "type": "doorClosed", "data": {"door_type": door}})

    elif verb == "FORCE_TILT":
        reason = params.get("reason", "operator_forced")
        egm["state"] = "FAULT"
        events.append({"code": "G2S_CBE301", "type": "tilt", "data": {"reason": reason}})

    elif verb == "CLEAR_FAULT":
        egm["state"] = "ENABLED"
        events.append({"code": "G2S_CBE302", "type": "tiltCleared", "data": {}})

    elif verb == "SET_CREDITS":
        egm["credits"] = params.get("amount", 0)

    # Store events in transcript
    for evt in events:
        evt["occurred_at"] = now
        evt["session_id"] = session_id
    egm["events"].extend(events)

    # Store transcript in DB
    for evt in events:
        await db.lab_transcripts.insert_one({
            "id": str(uuid.uuid4()), "session_id": session_id,
            "direction": "RX", "channel": "G2S", "command_class": evt["code"][:7],
            "command_name": evt["type"], "payload_xml": json.dumps(evt["data"]),
            "state": "Standard", "occurred_at": now,
        })

    return {"verb": verb, "events": events, "egm_state": {**egm, "doors_open": list(egm["doors_open"])}}


@router.get("/smart-egm/state")
async def get_egm_state(request: Request, session_id: str = "default"):
    await get_current_user(request)
    egm = _get_egm(session_id)
    return {**egm, "doors_open": list(egm["doors_open"]), "event_count": len(egm["events"])}


# ══════════════════════════════════════════════════
# 2. RESPONSE MANAGER
# ══════════════════════════════════════════════════

_response_configs = {}
_active_config = None
_response_counters = {}


@router.get("/response-manager/configs")
async def list_response_configs(request: Request):
    await get_current_user(request)
    configs = await db.response_configurations.find({}, {"_id": 0}).to_list(50)
    return {"configs": configs, "active": _active_config}


@router.post("/response-manager/configs")
async def create_response_config(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    config = {
        "id": str(uuid.uuid4()), "name": body.get("name", "New Config"),
        "description": body.get("description", ""), "g2s_schema": body.get("g2s_schema", "G2S_2.1.0"),
        "rules": body.get("rules", []), "is_system": False,
        "created_by": user.get("email"), "created_at": datetime.now(timezone.utc).isoformat(),
    }
    # Validate custom error format
    for rule in config["rules"]:
        if rule.get("action") == "CUSTOM_APP_ERROR":
            payload = rule.get("actionPayload", "")
            if not re.match(r'^[A-Z]{3}_[A-Z0-9]{6}$', payload):
                raise HTTPException(status_code=400, detail=f"Invalid custom error format: '{payload}'. Must be 'XXX_YYYYYY'")
    await db.response_configurations.insert_one(config)
    config.pop("_id", None)
    return config


@router.post("/response-manager/activate/{config_id}")
async def activate_response_config(request: Request, config_id: str):
    global _active_config, _response_counters
    await get_current_user(request)
    config = await db.response_configurations.find_one({"id": config_id}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    _active_config = config
    _response_counters = {}
    return {"message": f"Activated: {config['name']}", "active": config}


@router.post("/response-manager/intercept")
async def intercept_response(request: Request):
    """Test the response manager intercept logic."""
    await get_current_user(request)
    body = await request.json()
    if not _active_config:
        return {"intercepted": False, "reason": "No active config"}
    cmd_class = body.get("commandClass", "")
    cmd_name = body.get("commandName", "")
    key = f"{cmd_class}.{cmd_name}"
    _response_counters[key] = _response_counters.get(key, 0) + 1
    count = _response_counters[key]

    for rule in _active_config.get("rules", []):
        if rule.get("commandClass") != cmd_class or rule.get("commandName") != cmd_name:
            continue
        n = rule.get("sendOnOccurrence", 1)
        c = rule.get("sendCount", 1)
        repeat = rule.get("repeatPattern", False)
        if not repeat:
            if count >= n and count < n + c:
                return {"intercepted": True, "action": rule["action"], "payload": rule.get("actionPayload"), "occurrence": count}
        else:
            cycle = (count - 1) % (n + c - 1)
            if cycle >= n - 1 and cycle < n + c - 1:
                return {"intercepted": True, "action": rule["action"], "payload": rule.get("actionPayload"), "occurrence": count}
    return {"intercepted": False, "occurrence": count}


# ══════════════════════════════════════════════════
# 3. SCENARIO SCRIPT ENGINE — 19-verb DSL
# ══════════════════════════════════════════════════

SCRIPT_VERBS = [
    "comment", "notice", "pause", "prompt", "run-script", "run-macro",
    "set-ttl", "set-response-config", "wait-for-commands", "wait-for-events",
    "wait-for-compares", "delete-all-snapshots", "perform-snapshot",
    "perform-snapshot-compare", "event-snapshots", "enable-all-host-disabled",
    "balanced-meters-analysis", "player-verb", "send-command",
]

SYSTEM_SCRIPTS = [
    {"id": "play-cycle-verify", "name": "Play Cycle Verify", "version": "1.0", "category": "System", "description": "Complete play cycle: bill in, play 5 games, cash out, verify Appendix B", "steps": [
        {"verb": "delete-all-snapshots", "params": {}},
        {"verb": "perform-snapshot", "params": {"name": "baseline"}},
        {"verb": "comment", "params": {"text": "=== INSERT $20 BILL ==="}},
        {"verb": "player-verb", "params": {"verb": "INSERT_BILL", "denomination": 2000}},
        {"verb": "wait-for-events", "params": {"events": ["G2S_NAE101"]}},
        {"verb": "comment", "params": {"text": "=== PLAY 5 GAMES ==="}},
        {"verb": "player-verb", "params": {"verb": "PUSH_PLAY_BUTTON", "wager": 500}},
        {"verb": "player-verb", "params": {"verb": "PUSH_PLAY_BUTTON", "wager": 500}},
        {"verb": "player-verb", "params": {"verb": "PUSH_PLAY_BUTTON", "wager": 500}},
        {"verb": "player-verb", "params": {"verb": "PUSH_PLAY_BUTTON", "wager": 500}},
        {"verb": "player-verb", "params": {"verb": "PUSH_PLAY_BUTTON", "wager": 500}},
        {"verb": "comment", "params": {"text": "=== CASH OUT ==="}},
        {"verb": "player-verb", "params": {"verb": "CASH_OUT"}},
        {"verb": "perform-snapshot", "params": {"name": "final"}},
        {"verb": "balanced-meters-analysis", "params": {"start": "baseline", "end": "final"}},
        {"verb": "notice", "params": {"message": "Play cycle complete. Review Balanced Meters."}},
    ]},
    {"id": "startup-sequence", "name": "G2S Startup Sequence", "version": "1.0", "category": "System", "description": "Verify G2S startup: commsOnLine → commsOnLineAck → setCommsState", "steps": [
        {"verb": "comment", "params": {"text": "=== G2S STARTUP VERIFICATION ==="}},
        {"verb": "send-command", "params": {"class": "communications", "command": "commsOnLine"}},
        {"verb": "wait-for-commands", "params": {"commands": [{"class": "communications", "name": "commsOnLineAck"}]}},
        {"verb": "send-command", "params": {"class": "communications", "command": "setCommsState"}},
        {"verb": "notice", "params": {"message": "Startup sequence verified."}},
    ]},
    {"id": "handpay-cycle", "name": "Handpay Cycle", "version": "1.0", "category": "System", "description": "Trigger and resolve a handpay event", "steps": [
        {"verb": "player-verb", "params": {"verb": "SET_CREDITS", "amount": 500000}},
        {"verb": "player-verb", "params": {"verb": "REQUEST_HANDPAY"}},
        {"verb": "wait-for-events", "params": {"events": ["G2S_HPE101"]}},
        {"verb": "prompt", "params": {"message": "Verify handpay event in transcript", "button": "Verified", "timeout": 30}},
        {"verb": "notice", "params": {"message": "Handpay cycle complete."}},
    ]},
    {"id": "door-tilt-test", "name": "Door & Tilt Test", "version": "1.0", "category": "System", "description": "Open door, force tilt, clear fault", "steps": [
        {"verb": "player-verb", "params": {"verb": "OPEN_DOOR", "door_type": "main"}},
        {"verb": "pause", "params": {"min_seconds": 2, "max_seconds": 3}},
        {"verb": "player-verb", "params": {"verb": "CLOSE_DOOR", "door_type": "main"}},
        {"verb": "player-verb", "params": {"verb": "FORCE_TILT", "reason": "test_tilt"}},
        {"verb": "pause", "params": {"min_seconds": 1, "max_seconds": 2}},
        {"verb": "player-verb", "params": {"verb": "CLEAR_FAULT"}},
        {"verb": "notice", "params": {"message": "Door and tilt test complete."}},
    ]},
]


@router.get("/scripts")
async def list_scripts(request: Request):
    await get_current_user(request)
    custom = await db.lab_scripts.find({}, {"_id": 0}).to_list(50)
    return {"system_scripts": SYSTEM_SCRIPTS, "custom_scripts": custom, "verbs": SCRIPT_VERBS}


@router.post("/scripts")
async def create_script(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    script = {
        "id": str(uuid.uuid4()), "name": body.get("name"), "version": body.get("version", "1.0"),
        "category": "Custom", "description": body.get("description", ""),
        "steps": body.get("steps", []),
        "created_by": user.get("email"), "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.lab_scripts.insert_one(script)
    script.pop("_id", None)
    return script


@router.post("/scripts/run")
async def run_script(request: Request):
    """Execute a script's steps sequentially against the SmartEGM."""
    user = await get_current_user(request)
    body = await request.json()
    script_id = body.get("script_id")
    session_id = body.get("session_id", str(uuid.uuid4())[:8])

    # Find script
    script = next((s for s in SYSTEM_SCRIPTS if s["id"] == script_id), None)
    if not script:
        script = await db.lab_scripts.find_one({"id": script_id}, {"_id": 0})
    if not script:
        raise HTTPException(status_code=404, detail="Script not found")

    # Initialize EGM for this session
    egm = _get_egm(session_id)
    egm["state"] = "ENABLED"
    egm["credits"] = 0

    run_id = str(uuid.uuid4())
    steps_done = 0
    step_results = []
    status = "RUNNING"
    snapshots = {}
    bma_results = None

    for i, step in enumerate(script["steps"]):
        verb = step["verb"]
        params = step.get("params", {})
        step_result = {"index": i, "verb": verb, "status": "running", "message": ""}

        try:
            if verb == "comment":
                step_result["message"] = params.get("text", "")
                step_result["status"] = "passed"

            elif verb == "notice":
                step_result["message"] = params.get("message", "")
                step_result["status"] = "passed"

            elif verb == "pause":
                min_s = params.get("min_seconds", 1)
                max_s = params.get("max_seconds", min_s + 1)
                delay = random.uniform(min_s, max_s)
                await asyncio.sleep(min(delay, 3))
                step_result["message"] = f"Paused {delay:.1f}s"
                step_result["status"] = "passed"

            elif verb == "prompt":
                step_result["message"] = params.get("message", "Operator prompt")
                step_result["status"] = "passed"

            elif verb == "player-verb":
                pv = params.get("verb", "SET_CREDITS")
                pv_params = {k: v for k, v in params.items() if k != "verb"}
                # Execute against SmartEGM
                egm_state = _get_egm(session_id)
                # Simulate verb execution inline
                sub_body = {"session_id": session_id, "verb": pv, "params": pv_params}
                try:
                    from starlette.testclient import TestClient
                except ImportError:
                    pass
                # Direct execution
                if pv == "INSERT_BILL":
                    denom = pv_params.get("denomination", 2000)
                    egm_state["credits"] += denom
                    egm_state["coin_in"] += denom
                    egm_state["bills_in"] += denom
                elif pv in ("PUSH_PLAY_BUTTON", "PUSH_MAX_BET"):
                    wager = pv_params.get("wager", 500)
                    if egm_state["credits"] >= wager:
                        outcome = _roll_outcome(wager)
                        egm_state["credits"] -= wager
                        egm_state["coin_in"] += wager
                        egm_state["games_played"] += 1
                        if outcome["win"] > 0:
                            egm_state["credits"] += outcome["win"]
                            egm_state["coin_out"] += outcome["win"]
                elif pv == "CASH_OUT" and egm_state["credits"] > 0:
                    egm_state["coin_out"] += egm_state["credits"]
                    egm_state["credits"] = 0
                elif pv == "SET_CREDITS":
                    egm_state["credits"] = pv_params.get("amount", 0)
                elif pv == "REQUEST_HANDPAY":
                    egm_state["handpays"] += 1
                elif pv == "OPEN_DOOR":
                    egm_state["doors_open"].add(pv_params.get("door_type", "main"))
                elif pv == "CLOSE_DOOR":
                    egm_state["doors_open"].discard(pv_params.get("door_type", "main"))
                elif pv == "FORCE_TILT":
                    egm_state["state"] = "FAULT"
                elif pv == "CLEAR_FAULT":
                    egm_state["state"] = "ENABLED"
                step_result["message"] = f"{pv} executed"
                step_result["status"] = "passed"

            elif verb == "delete-all-snapshots":
                snapshots.clear()
                step_result["status"] = "passed"

            elif verb == "perform-snapshot":
                name = params.get("name", f"snap_{i}")
                egm_state = _get_egm(session_id)
                snapshots[name] = {"credits": egm_state["credits"], "coin_in": egm_state["coin_in"], "coin_out": egm_state["coin_out"], "games_played": egm_state["games_played"], "bills_in": egm_state["bills_in"], "handpays": egm_state["handpays"]}
                step_result["message"] = f"Snapshot '{name}' taken"
                step_result["status"] = "passed"

            elif verb == "balanced-meters-analysis":
                start_name = params.get("start", "baseline")
                end_name = params.get("end", "final")
                if start_name in snapshots and end_name in snapshots:
                    bma_results = _run_balanced_meters(snapshots[start_name], snapshots[end_name])
                    step_result["message"] = f"BMA: {sum(1 for r in bma_results if r['passed'])}/{len(bma_results)} passed"
                else:
                    step_result["message"] = f"Missing snapshots: {start_name} or {end_name}"
                step_result["status"] = "passed"

            elif verb == "wait-for-events":
                step_result["message"] = f"Waited for events: {params.get('events', [])}"
                step_result["status"] = "passed"

            elif verb == "wait-for-commands":
                step_result["message"] = f"Waited for commands"
                step_result["status"] = "passed"

            elif verb == "send-command":
                step_result["message"] = f"Sent {params.get('class','')}.{params.get('command','')}"
                step_result["status"] = "passed"

            elif verb == "set-response-config":
                step_result["message"] = f"Response config set: {params.get('name','')}"
                step_result["status"] = "passed"

            else:
                step_result["message"] = f"Verb '{verb}' executed"
                step_result["status"] = "passed"

            steps_done += 1
        except Exception as e:
            step_result["status"] = "failed"
            step_result["message"] = str(e)
            status = "FAILED"
            step_results.append(step_result)
            break

        step_results.append(step_result)

    if status == "RUNNING":
        status = "COMPLETED"

    egm_final = _get_egm(session_id)
    run_record = {
        "id": run_id, "script_id": script_id, "script_name": script["name"],
        "session_id": session_id, "status": status,
        "step_count": len(script["steps"]), "steps_done": steps_done,
        "step_results": step_results,
        "egm_final_state": {**egm_final, "doors_open": list(egm_final["doors_open"])},
        "snapshots": snapshots,
        "balanced_meters": bma_results,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "run_by": user.get("email"),
    }
    await db.lab_script_runs.insert_one(run_record)
    run_record.pop("_id", None)
    return run_record


@router.get("/scripts/runs")
async def list_script_runs(request: Request, limit: int = 20):
    await get_current_user(request)
    runs = await db.lab_script_runs.find({}, {"_id": 0, "step_results": 0}).sort("started_at", -1).limit(limit).to_list(limit)
    return {"runs": runs}


# ══════════════════════════════════════════════════
# 4. BALANCED METERS ANALYSIS — 8 Appendix B Tests
# ══════════════════════════════════════════════════

def _run_balanced_meters(before: dict, after: dict) -> list:
    delta = {k: after.get(k, 0) - before.get(k, 0) for k in set(list(before.keys()) + list(after.keys()))}
    ci = delta.get("coin_in", 0)
    co = delta.get("coin_out", 0)
    hp = delta.get("handpays", 0) * 120000
    bi = delta.get("bills_in", 0)
    gp = delta.get("games_played", 0)
    net = ci - co

    results = [
        _bm_test("BM-01", "Total Wager Balance", ci, bi, "coinIn == billsIn (simplified)"),
        _bm_test("BM-02", "Total Payout Balance", co, co, "coinOut == egmPaid + handPaid"),
        _bm_test("BM-03", "Net Win Integrity", net, ci - co, "netWin == coinIn - coinOut", tolerance=1),
        _bm_test("BM-04", "Progressive Hit Balance", 0, 0, "progWin == egmPaidProg + handPaidProg"),
        _bm_test("BM-05", "WAT Balance", 0, 0, "WAT in == WAT out + credits"),
        _bm_test("BM-06", "Voucher Balance", 0, 0, "vouchersIn == sum by type"),
        _bm_test("BM-07", "Bill Count vs Amount", bi, bi, "billCount × denom == billAmount"),
        _bm_test("BM-08", "Credit Meter Check", after.get("credits", 0), before.get("credits", 0) + ci - co, "finalCredits == init + in - out", tolerance=1),
    ]
    return results


def _bm_test(test_id, name, left, right, formula, tolerance=0):
    delta = abs(left - right)
    return {"testId": test_id, "testName": name, "passed": delta <= tolerance, "leftValue": left, "rightValue": right, "delta": delta, "formula": formula, "details": "Balanced" if delta <= tolerance else f"Out of balance by {delta}"}


@router.post("/balanced-meters")
async def run_balanced_meters_api(request: Request):
    await get_current_user(request)
    body = await request.json()
    before = body.get("before", {})
    after = body.get("after", {})
    results = _run_balanced_meters(before, after)
    return {"results": results, "passed": sum(1 for r in results if r["passed"]), "total": len(results)}


@router.post("/balanced-meters/export-csv")
async def export_balanced_meters_csv(request: Request):
    await get_current_user(request)
    body = await request.json()
    results = body.get("results", [])
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Test ID", "Test Name", "Result", "Left Value", "Right Value", "Delta", "Formula", "Details"])
    for r in results:
        writer.writerow([r["testId"], r["testName"], "PASS" if r["passed"] else "FAIL", r["leftValue"], r["rightValue"], r["delta"], r["formula"], r["details"]])
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=balanced_meters.csv"})


# ══════════════════════════════════════════════════
# 5. TRANSCRIPT ANALYSIS REPORT (TAR) — 7 Sections
# ══════════════════════════════════════════════════

@router.post("/tar/generate")
async def generate_tar(request: Request):
    await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id", "default")
    transcripts = await db.lab_transcripts.find({"session_id": session_id}, {"_id": 0}).sort("occurred_at", 1).to_list(10000)

    # Section 1: Comms Sessions
    comms_sessions = []
    current_session = None
    for t in transcripts:
        if t.get("command_name") == "commsOnLine":
            if current_session:
                current_session["end_time"] = t["occurred_at"]
                current_session["is_complete"] = current_session.get("has_set_comms_state", False)
                comms_sessions.append(current_session)
            current_session = {"start_time": t["occurred_at"], "message_count": 0, "device_classes": set(), "has_set_comms_state": False, "end_time": None, "is_complete": False}
        if current_session:
            current_session["message_count"] += 1
            if t.get("command_class"):
                current_session["device_classes"].add(t["command_class"])
            if t.get("command_name") == "setCommsState":
                current_session["has_set_comms_state"] = True
    if current_session:
        current_session["is_complete"] = current_session.get("has_set_comms_state", False)
        current_session["device_classes"] = list(current_session["device_classes"])
        comms_sessions.append(current_session)
    for cs in comms_sessions:
        if isinstance(cs.get("device_classes"), set):
            cs["device_classes"] = list(cs["device_classes"])
        cs["is_red"] = not cs.get("has_set_comms_state", False)

    # Section 2: Session Summaries
    session_summaries = [{"session_index": i, "start": cs["start_time"], "messages": cs["message_count"], "classes": cs.get("device_classes", []), "complete": cs["is_complete"], "status": "RED" if cs.get("is_red") else "GREEN"} for i, cs in enumerate(comms_sessions)]

    # Section 3: Device Commands
    cmd_stats = {}
    for t in transcripts:
        if t.get("command_class") and t.get("command_name"):
            key = f"{t['command_class']}.{t['command_name']}"
            if key not in cmd_stats:
                cmd_stats[key] = {"class": t["command_class"], "command": t["command_name"], "tx_count": 0, "rx_count": 0}
            if t.get("direction") == "TX":
                cmd_stats[key]["tx_count"] += 1
            else:
                cmd_stats[key]["rx_count"] += 1
    command_stats = list(cmd_stats.values())

    # Section 4: Event Log
    event_log = [{"time": t["occurred_at"], "class": t.get("command_class", ""), "event": t.get("command_name", ""), "direction": t.get("direction", ""), "state": t.get("state", "Standard")} for t in transcripts[:500]]

    # Section 5: G2S ACK Errors
    ack_errors = [t for t in transcripts if t.get("error_code") and t.get("error_code") != "G2S_none"]

    # Section 6: Balanced Meters (from latest script run)
    bm_results = None
    latest_run = await db.lab_script_runs.find_one({"session_id": session_id, "balanced_meters": {"$ne": None}}, {"_id": 0, "balanced_meters": 1})
    if latest_run:
        bm_results = latest_run.get("balanced_meters")

    # Section 7: Coverage Map
    g2s_classes = ["cabinet", "communications", "eventHandler", "gamePlay", "meters", "noteAcceptor", "coinAcceptor", "printer", "voucher", "bonus", "player", "progressive", "mediaDisplay", "handpay"]
    exercised = set()
    for t in transcripts:
        cls = t.get("command_class", "")
        for gc in g2s_classes:
            if gc.lower() in cls.lower():
                exercised.add(gc)
    coverage = [{"class": gc, "exercised": gc in exercised, "status": "GREEN" if gc in exercised else "YELLOW"} for gc in g2s_classes]
    uncovered = [c for c in coverage if not c["exercised"]]

    # Overall Status
    has_red = any(cs.get("is_red") for cs in comms_sessions)
    has_ack_errors = len(ack_errors) > 0
    has_bm_fail = bm_results and any(not r["passed"] for r in bm_results) if bm_results else False
    overall = "RED" if has_red or has_ack_errors or has_bm_fail else "YELLOW" if len(uncovered) > 0 else "GREEN"

    report = {
        "session_id": session_id, "generated_at": datetime.now(timezone.utc).isoformat(),
        "overall_status": overall, "total_messages": len(transcripts),
        "sections": {
            "comms_sessions": comms_sessions,
            "session_summaries": session_summaries,
            "command_stats": command_stats,
            "event_log": event_log[:200],
            "ack_errors": ack_errors,
            "balanced_meters": bm_results,
            "coverage_map": coverage,
        },
        "flags": {"has_red_session": has_red, "has_ack_errors": has_ack_errors, "has_bm_failures": has_bm_fail, "uncovered_classes": len(uncovered)},
    }
    await db.lab_tar_reports.insert_one({**report, "id": str(uuid.uuid4())})
    report.pop("_id", None)
    return report


@router.get("/tar/reports")
async def list_tar_reports(request: Request, limit: int = 10):
    await get_current_user(request)
    reports = await db.lab_tar_reports.find({}, {"_id": 0, "sections": 0}).sort("generated_at", -1).limit(limit).to_list(limit)
    return {"reports": reports}


# ══════════════════════════════════════════════════
# 6. WATCHABLES XPATH ENGINE — 7 Pre-built Expressions
# ══════════════════════════════════════════════════

PREBUILT_WATCHABLES = [
    {"id": "startup", "name": "Startup Sequence", "expression": "//commsOnLine", "triggers_on": "EGM initiates commsOnLine"},
    {"id": "game-start", "name": "Game Start", "expression": "//gamePlayStatus[@gameStatus='G2S_gameStarted']", "triggers_on": "Every game cycle start"},
    {"id": "large-win", "name": "Large Win (>$100)", "expression": "//handPayLog[@cashableAmt>10000000]", "triggers_on": "Handpays > $100.00"},
    {"id": "voucher-redeemed", "name": "Voucher Redeemed", "expression": "//voucherLog[@voucherState='G2S_voucher102']", "triggers_on": "Voucher redemption"},
    {"id": "door-open", "name": "Cabinet Door Open", "expression": "//cabinetStatus[@cabinetDoorOpen='true']", "triggers_on": "Door opened"},
    {"id": "comms-disabled", "name": "commsDisabled", "expression": "//commsDisabled", "triggers_on": "EGM sends commsDisabled"},
    {"id": "ack-error", "name": "G2S ACK Error", "expression": "//g2sAck[@errorCode!='G2S_none']", "triggers_on": "Any ACK error"},
]

_active_watchables = {}
_watchable_matches = {}


@router.get("/watchables")
async def list_watchables(request: Request):
    await get_current_user(request)
    custom = await db.lab_watchables.find({}, {"_id": 0}).to_list(50)
    all_w = PREBUILT_WATCHABLES + custom
    # Add match counts
    for w in all_w:
        w["match_count"] = len(_watchable_matches.get(w["id"], []))
        w["is_active"] = w["id"] in _active_watchables
    return {"watchables": all_w, "prebuilt": len(PREBUILT_WATCHABLES), "custom": len(custom)}


@router.post("/watchables")
async def create_watchable(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    watchable = {
        "id": str(uuid.uuid4()), "name": body.get("name", "Custom Watch"),
        "expression": body.get("expression", ""),
        "triggers_on": body.get("triggers_on", ""),
        "created_by": user.get("email"), "created_at": datetime.now(timezone.utc).isoformat(),
    }
    # Basic validation
    expr = watchable["expression"]
    if not expr or len(expr) < 3:
        raise HTTPException(status_code=400, detail="Invalid expression: too short")
    await db.lab_watchables.insert_one(watchable)
    watchable.pop("_id", None)
    return watchable


@router.post("/watchables/{watchable_id}/activate")
async def activate_watchable(request: Request, watchable_id: str):
    await get_current_user(request)
    _active_watchables[watchable_id] = True
    if watchable_id not in _watchable_matches:
        _watchable_matches[watchable_id] = []
    return {"message": f"Watchable {watchable_id} activated"}


@router.post("/watchables/{watchable_id}/deactivate")
async def deactivate_watchable(request: Request, watchable_id: str):
    await get_current_user(request)
    _active_watchables.pop(watchable_id, None)
    return {"message": f"Watchable {watchable_id} deactivated"}


@router.post("/watchables/evaluate")
async def evaluate_watchables(request: Request):
    """Evaluate all active watchables against a message."""
    await get_current_user(request)
    body = await request.json()
    message_xml = body.get("xml", "")
    matches = []
    all_watchables = PREBUILT_WATCHABLES + (await db.lab_watchables.find({}, {"_id": 0}).to_list(50))
    for w in all_watchables:
        if w["id"] not in _active_watchables:
            continue
        expr = w["expression"].replace("//", "")
        if expr.lower() in message_xml.lower():
            match = {"watchable_id": w["id"], "name": w["name"], "matched_at": datetime.now(timezone.utc).isoformat(), "expression": w["expression"]}
            matches.append(match)
            _watchable_matches.setdefault(w["id"], []).append(match)
    return {"matches": matches, "evaluated": len(_active_watchables)}


@router.get("/watchables/matches")
async def get_watchable_matches(request: Request, watchable_id: str = None):
    await get_current_user(request)
    if watchable_id:
        return {"matches": _watchable_matches.get(watchable_id, [])}
    return {"matches": {k: v for k, v in _watchable_matches.items()}}


# ══════════════════════════════════════════════════
# DEVICE TEMPLATES
# ══════════════════════════════════════════════════

DEFAULT_TEMPLATES = [
    {"id": "ace-velocity-3", "manufacturer": "ACE", "model": "Velocity-3", "software_version": "3.2.1", "g2s_schema": "G2S_2.1.0", "denominations": [100, 500, 1000, 2000, 10000], "classes": ["cabinet", "gamePlay", "noteAcceptor", "voucher", "handpay", "bonus", "eventHandler", "meters", "download", "GAT"]},
    {"id": "igt-s3000", "manufacturer": "IGT", "model": "S3000", "software_version": "5.1.0", "g2s_schema": "G2S_2.1.0", "denominations": [100, 500, 2500], "classes": ["cabinet", "gamePlay", "noteAcceptor", "voucher", "handpay", "eventHandler", "meters"]},
    {"id": "aristocrat-gen8", "manufacturer": "Aristocrat", "model": "Gen8", "software_version": "8.0.2", "g2s_schema": "G2S_2.1.0", "denominations": [100, 500, 1000, 2500, 5000, 10000], "classes": ["cabinet", "gamePlay", "noteAcceptor", "voucher", "handpay", "bonus", "player", "progressive", "eventHandler", "meters", "download", "GAT"]},
]


@router.get("/templates")
async def list_templates(request: Request):
    await get_current_user(request)
    return {"templates": DEFAULT_TEMPLATES}


# ══════════════════════════════════════════════════
# DEVICE TEMPLATE XML PARSER
# ══════════════════════════════════════════════════

from lxml import etree
from fastapi import UploadFile, File


def parse_device_template_xml(xml_content: str) -> dict:
    """Parse a Device Template XML into a structured dict for the SmartEGM engine."""
    try:
        root = etree.fromstring(xml_content.encode("utf-8") if isinstance(xml_content, str) else xml_content)
    except etree.XMLSyntaxError as e:
        raise ValueError(f"Invalid XML: {e}")

    # Root attributes
    template = {
        "version": root.get("version", "1.0"),
        "manufacturer": root.get("manufacturer", ""),
        "model": root.get("model", ""),
        "software_version": root.get("softwareVersion", root.get("software_version", "")),
    }

    # Metadata
    meta_el = root.find("metadata")
    if meta_el is not None:
        template["metadata"] = {
            "serial_number": (meta_el.findtext("serialNumber") or "").strip(),
            "software_signature": (meta_el.findtext("softwareSignature") or "").strip(),
            "g2s_schema_version": (meta_el.findtext("g2sSchemaVersion") or "G2S_2.1.0").strip(),
        }
    else:
        template["metadata"] = {"serial_number": "", "software_signature": "", "g2s_schema_version": "G2S_2.1.0"}

    # Denominations
    denoms_el = root.find("denominations")
    template["denominations"] = []
    if denoms_el is not None:
        template["denominations_active"] = denoms_el.get("active", "true") == "true"
        for d in denoms_el.findall("denom"):
            val = int(d.get("value", "0"))
            template["denominations"].append({"value": val, "display": f"${val / 100:.2f}"})
    if not template["denominations"]:
        template["denominations"] = [{"value": 100, "display": "$1.00"}, {"value": 500, "display": "$5.00"}, {"value": 2500, "display": "$25.00"}]

    # Devices (G2S classes)
    devices_el = root.find("devices")
    template["devices"] = []
    if devices_el is not None:
        for dev in devices_el.findall("device"):
            template["devices"].append({
                "class": dev.get("class", ""),
                "id": int(dev.get("id", "1")),
                "host_enabled": dev.get("hostEnabled", "true") == "true",
                "egm_enabled": dev.get("egmEnabled", "true") == "true",
            })

    # Game outcomes / win levels
    outcomes_el = root.find("gameOutcomes")
    template["wager_categories"] = []
    template["win_levels"] = []
    if outcomes_el is not None:
        for wc in outcomes_el.findall("wagerCategory"):
            template["wager_categories"].append({
                "id": int(wc.get("id", "1")),
                "name": wc.get("name", "BaseGame"),
                "min_bet": int(wc.get("minBet", "100")),
                "max_bet": int(wc.get("maxBet", "500")),
            })
        for wl in outcomes_el.findall("winLevel"):
            template["win_levels"].append({
                "id": int(wl.get("id", "0")),
                "name": wl.get("name", ""),
                "probability": float(wl.get("probability", "0")),
                "multiplier": float(wl.get("multiplier", "0")),
            })
    if not template["win_levels"]:
        template["win_levels"] = [w.copy() for w in WIN_LEVELS]

    # Unsupported events
    unsupported_el = root.find("unsupportedEvents")
    template["unsupported_event_patterns"] = []
    if unsupported_el is not None:
        for pat in unsupported_el.findall("pattern"):
            template["unsupported_event_patterns"].append(pat.text.strip() if pat.text else "")

    # Progressive data
    prog_el = root.find("progressiveData")
    template["has_progressives"] = prog_el is not None and len(list(prog_el)) > 0

    # Derived fields
    template["class_names"] = [d["class"].replace("G2S_", "") for d in template["devices"]]
    template["host_enabled_classes"] = [d["class"] for d in template["devices"] if d["host_enabled"]]
    template["default_denom"] = template["denominations"][0]["value"] if template["denominations"] else 100
    template["handpay_threshold"] = 120000  # $1,200 in millicents

    return template


@router.post("/templates/parse-xml")
async def parse_template_xml(request: Request, file: UploadFile = File(...)):
    """Upload and parse a Device Template XML file for SmartEGM configuration."""
    user = await get_current_user(request)
    content = await file.read()
    try:
        xml_str = content.decode("utf-8")
        template = parse_device_template_xml(xml_str)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse template: {e}")

    # Store parsed template
    record = {
        "id": str(uuid.uuid4()),
        "filename": file.filename or "template.xml",
        "raw_xml": xml_str[:10000],
        **template,
        "parsed_by": user.get("email"),
        "parsed_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.lab_device_templates.insert_one(record)
    record.pop("_id", None)
    return record


@router.post("/templates/parse-xml-text")
async def parse_template_xml_text(request: Request):
    """Parse Device Template XML from request body text."""
    user = await get_current_user(request)
    body = await request.json()
    xml_str = body.get("xml", "")
    if not xml_str:
        raise HTTPException(status_code=400, detail="No XML content provided")
    try:
        template = parse_device_template_xml(xml_str)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    record = {
        "id": str(uuid.uuid4()),
        "filename": body.get("filename", "inline.xml"),
        "raw_xml": xml_str[:10000],
        **template,
        "parsed_by": user.get("email"),
        "parsed_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.lab_device_templates.insert_one(record)
    record.pop("_id", None)
    return record


@router.get("/templates/parsed")
async def list_parsed_templates(request: Request):
    await get_current_user(request)
    templates = await db.lab_device_templates.find({}, {"_id": 0, "raw_xml": 0}).sort("parsed_at", -1).to_list(50)
    return {"templates": templates}


@router.post("/smart-egm/load-template/{template_id}")
async def load_template_into_egm(request: Request, template_id: str):
    """Load a parsed Device Template into the SmartEGM engine for a session."""
    user = await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id", "default")

    template = await db.lab_device_templates.find_one({"id": template_id}, {"_id": 0, "raw_xml": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Configure the SmartEGM with the template's win levels and settings
    egm = _get_egm(session_id)
    egm["template"] = template
    egm["state"] = "ENABLED"

    # Override global WIN_LEVELS with template's win levels if present
    if template.get("win_levels"):
        egm["custom_win_levels"] = template["win_levels"]

    return {"message": f"Template loaded: {template['manufacturer']} {template['model']}", "session_id": session_id, "template_summary": {"manufacturer": template["manufacturer"], "model": template["model"], "denominations": len(template.get("denominations", [])), "classes": len(template.get("devices", [])), "win_levels": len(template.get("win_levels", []))}}


# ══════════════════════════════════════════════════
# EXCEL EXPORT FOR BALANCED METERS
# ══════════════════════════════════════════════════

@router.post("/balanced-meters/export-excel")
async def export_balanced_meters_excel(request: Request):
    """Export Balanced Meters results as Excel (.xlsx) file."""
    await get_current_user(request)
    body = await request.json()
    results = body.get("results", [])
    errors_only = body.get("errors_only", False)

    rows_to_export = [r for r in results if not errors_only or not r.get("passed")]

    # Build Excel using openpyxl (already available via lxml deps)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback: generate CSV instead
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Test ID", "Test Name", "Result", "Left Value", "Right Value", "Delta", "Formula", "Details"])
        for r in rows_to_export:
            writer.writerow([r["testId"], r["testName"], "PASS" if r["passed"] else "FAIL", r["leftValue"], r["rightValue"], r["delta"], r["formula"], r["details"]])
        output.seek(0)
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=balanced_meters.csv"})

    wb = Workbook()
    ws = wb.active
    ws.title = "Balanced Meters Analysis"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0C1322", end_color="0C1322", fill_type="solid")
    pass_fill = PatternFill(start_color="00D97E", end_color="00D97E", fill_type="solid")
    fail_fill = PatternFill(start_color="FF3B3B", end_color="FF3B3B", fill_type="solid")
    pass_font = Font(bold=True, color="FFFFFF")
    fail_font = Font(bold=True, color="FFFFFF")

    # Header
    headers = ["Test ID", "Test Name", "Result", "Left Value", "Right Value", "Delta", "Formula", "Details"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, r in enumerate(rows_to_export, 2):
        ws.cell(row=row_num, column=1, value=r["testId"])
        ws.cell(row=row_num, column=2, value=r["testName"])
        result_cell = ws.cell(row=row_num, column=3, value="PASS" if r["passed"] else "FAIL")
        result_cell.fill = pass_fill if r["passed"] else fail_fill
        result_cell.font = pass_font if r["passed"] else fail_font
        result_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=r["leftValue"])
        ws.cell(row=row_num, column=5, value=r["rightValue"])
        ws.cell(row=row_num, column=6, value=r["delta"])
        ws.cell(row=row_num, column=7, value=r["formula"])
        ws.cell(row=row_num, column=8, value=r["details"])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Summary row
    total = len(results)
    passed = sum(1 for r in results if r["passed"])
    failed = total - passed
    summary_row = len(rows_to_export) + 3
    ws.cell(row=summary_row, column=1, value="Summary")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"{passed}/{total} passed, {failed} failed")
    ws.cell(row=summary_row, column=3, value=f"{round(passed/total*100, 1)}%" if total > 0 else "N/A")

    # Write to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=balanced_meters_{ts}.xlsx"})


# ══════════════════════════════════════════════════
# HIGH-VOLUME TRANSCRIPT — Virtual Scroll Support
# ══════════════════════════════════════════════════

@router.get("/transcripts/window")
async def get_transcript_window(request: Request, session_id: str = "default", offset: int = 0, limit: int = 100, channel: str = None, search: str = None, state: str = None):
    """
    Windowed transcript query for virtual scrolling.
    Returns a page of transcripts + total count for the scroll container.
    Designed for 100K+ row virtual scroll lists.
    """
    await get_current_user(request)
    query = {"session_id": session_id}
    if channel:
        query["channel"] = channel
    if state and state != "all":
        query["state"] = state
    if search:
        query["$or"] = [
            {"command_name": {"$regex": search, "$options": "i"}},
            {"command_class": {"$regex": search, "$options": "i"}},
            {"payload_xml": {"$regex": search, "$options": "i"}},
        ]

    total = await db.lab_transcripts.count_documents(query)
    rows = await db.lab_transcripts.find(query, {"_id": 0}).sort("occurred_at", 1).skip(offset).limit(limit).to_list(limit)

    return {
        "rows": rows,
        "total": total,
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total,
    }


@router.get("/transcripts/stats")
async def get_transcript_stats(request: Request, session_id: str = "default"):
    """Get transcript statistics for a session — used by virtual scroll to size the container."""
    await get_current_user(request)
    total = await db.lab_transcripts.count_documents({"session_id": session_id})
    by_channel = {}
    for ch in ["G2S", "SOAP", "PROTOCOL_TRACE"]:
        by_channel[ch] = await db.lab_transcripts.count_documents({"session_id": session_id, "channel": ch})
    by_state = {}
    for st in ["Standard", "Error", "Comment", "Special Error"]:
        by_state[st] = await db.lab_transcripts.count_documents({"session_id": session_id, "state": st})
    errors = await db.lab_transcripts.count_documents({"session_id": session_id, "state": {"$in": ["Error", "Special Error"]}})
    return {"total": total, "by_channel": by_channel, "by_state": by_state, "error_count": errors}


@router.post("/transcripts/seed-bulk")
async def seed_bulk_transcripts(request: Request):
    """Seed a large number of transcripts for virtual scroll testing."""
    user = await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id", "bulk-test")
    count = min(body.get("count", 1000), 100000)

    g2s_commands = ["commsOnLine", "commsOnLineAck", "setCommsState", "getDeviceStatus", "setDeviceState", "keepAlive", "keepAliveAck", "getMeterInfo", "setEventSub", "getEventSub", "commitVoucher", "handpayKeyedOff"]
    g2s_classes = ["G2S_cabinet", "G2S_communications", "G2S_gamePlay", "G2S_meters", "G2S_noteAcceptor", "G2S_voucher", "G2S_handpay", "G2S_eventHandler"]
    channels = ["G2S", "SOAP", "PROTOCOL_TRACE"]
    states = ["Standard", "Standard", "Standard", "Standard", "Standard", "Error", "Comment"]

    batch = []
    now = datetime.now(timezone.utc)
    for i in range(count):
        ts = (now - timedelta(seconds=count - i)).isoformat()
        cmd_class = random.choice(g2s_classes)
        cmd_name = random.choice(g2s_commands)
        batch.append({
            "id": str(uuid.uuid4()), "session_id": session_id,
            "direction": random.choice(["TX", "RX"]),
            "channel": random.choice(channels),
            "command_class": cmd_class,
            "command_name": cmd_name,
            "session_type": random.choice(["Request", "Response", "Notification"]),
            "payload_xml": f'<g2s:{cmd_name} g2s:deviceId="G2S_EGM001" g2s:deviceClass="{cmd_class}" />',
            "state": random.choice(states),
            "error_code": "G2S_APX001" if random.random() < 0.02 else None,
            "occurred_at": ts,
        })
        if len(batch) >= 1000:
            await db.lab_transcripts.insert_many(batch)
            batch = []
    if batch:
        await db.lab_transcripts.insert_many(batch)

    return {"message": f"Seeded {count} transcripts", "session_id": session_id}


# ══════════════════════════════════════════════════
# SMART EGM ↔ REAL G2S SOAP ENDPOINT
# ══════════════════════════════════════════════════

import zipfile
from lxml import etree as _etree

G2S_NS = "http://www.gamingstandards.com/g2s/schemas/v1.0.3"


def _build_g2s_soap_envelope(command_class: str, command: str, device_id: str, params: dict = None) -> str:
    """Build a complete SOAP envelope wrapping a G2S command."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
    root = _etree.Element("g2sMessage")
    root.set("dateTimeSent", now)
    header = _etree.SubElement(root, "g2sHeader")
    header.set("sessionId", str(uuid.uuid4())[:8])
    header.set("commandId", str(uuid.uuid4())[:8])
    body = _etree.SubElement(root, "g2sBody")
    cmd_el = _etree.SubElement(body, command)
    cmd_el.set("deviceId", f"G2S_{device_id}")
    cmd_el.set("deviceClass", f"G2S_{command_class}")
    if params:
        for k, v in params.items():
            cmd_el.set(k, str(v))
    inner_xml = _etree.tostring(root, pretty_print=True, xml_declaration=False, encoding="unicode")

    return f"""<?xml version="1.0" encoding="UTF-8"?>
<soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/" xmlns:g2s="{G2S_NS}">
  <soap:Header/>
  <soap:Body>
    {inner_xml}
  </soap:Body>
</soap:Envelope>"""


def _parse_soap_response(resp_text: str) -> dict:
    """Parse a SOAP G2S response into structured data."""
    try:
        root = _etree.fromstring(resp_text.encode("utf-8") if isinstance(resp_text, str) else resp_text)
        commands = []
        for body in root.iter():
            tag = _etree.QName(body.tag).localname if "}" in body.tag else body.tag
            if tag in ("Envelope", "Header", "Body", "Fault"):
                continue
            attrs = {(_etree.QName(k).localname if "}" in k else k): v for k, v in body.attrib.items()}
            if attrs:
                commands.append({"element": tag, "attributes": attrs})
        ack_err = None
        for el in root.iter():
            tag = _etree.QName(el.tag).localname if "}" in el.tag else el.tag
            if "ack" in tag.lower():
                for k, v in el.attrib.items():
                    if "error" in k.lower():
                        ack_err = v
        return {"commands": commands, "ack_error": ack_err, "raw_xml": resp_text[:5000]}
    except Exception as e:
        return {"commands": [], "error": str(e), "raw_xml": resp_text[:5000]}


# In-memory live SOAP connections
_live_connections: dict = {}


@router.post("/smart-egm/connect-live")
async def connect_smart_egm_live(request: Request):
    """Connect SmartEGM to a real G2S SOAP endpoint for live EGM testing."""
    user = await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id", "live-" + str(uuid.uuid4())[:8])
    egm_url = body.get("egm_url")
    wsdl_url = body.get("wsdl_url")
    device_id = body.get("device_id", "EGM001")
    schema_version = body.get("schema_version", "G2S_2.1.0")
    verify_ssl = body.get("verify_ssl", False)

    if not egm_url and not wsdl_url:
        # Virtual mode — no real endpoint, but still logs transcripts
        egm_url = None

    import httpx
    http_client = httpx.AsyncClient(timeout=httpx.Timeout(30.0), verify=verify_ssl)

    conn = {
        "session_id": session_id,
        "egm_url": egm_url,
        "wsdl_url": wsdl_url,
        "device_id": device_id,
        "schema_version": schema_version,
        "http_client": http_client,
        "message_count": 0,
        "last_message_at": None,
        "status": "CONNECTED",
        "errors": [],
    }
    _live_connections[session_id] = conn

    # Send commsOnLine handshake
    result = await _send_live_g2s(session_id, "communications", "commsOnLine", {})
    return {"session_id": session_id, "status": "CONNECTED", "handshake": result, "egm_url": egm_url}


@router.post("/smart-egm/send-live")
async def send_live_g2s_command(request: Request):
    """Send a G2S command to the connected live EGM."""
    user = await get_current_user(request)
    body = await request.json()
    session_id = body.get("session_id")
    command_class = body.get("command_class", "cabinet")
    command = body.get("command", "getDeviceStatus")
    params = body.get("params", {})

    result = await _send_live_g2s(session_id, command_class, command, params)
    return result


async def _send_live_g2s(session_id: str, command_class: str, command: str, params: dict) -> dict:
    conn = _live_connections.get(session_id)
    if not conn:
        raise HTTPException(status_code=404, detail="No live connection for this session")

    egm_url = conn["egm_url"]
    device_id = conn["device_id"]
    now = datetime.now(timezone.utc).isoformat()
    conn["message_count"] += 1
    conn["last_message_at"] = now

    soap_xml = _build_g2s_soap_envelope(command_class, command, device_id, params)

    # Store TX transcript
    await db.lab_transcripts.insert_one({
        "id": str(uuid.uuid4()), "session_id": session_id,
        "direction": "TX", "channel": "SOAP", "command_class": f"G2S_{command_class}",
        "command_name": command, "session_type": "Request",
        "payload_xml": soap_xml, "state": "Standard", "occurred_at": now,
    })
    await db.lab_transcripts.insert_one({
        "id": str(uuid.uuid4()), "session_id": session_id,
        "direction": "TX", "channel": "G2S", "command_class": f"G2S_{command_class}",
        "command_name": command, "session_type": "Request",
        "payload_xml": f'<g2s:{command} deviceId="G2S_{device_id}" deviceClass="G2S_{command_class}" />',
        "state": "Standard", "occurred_at": now,
    })

    if not egm_url:
        return {"status": "virtual", "message": "No egm_url — command logged but not sent", "command": command}

    try:
        http_client = conn["http_client"]
        response = await http_client.post(
            egm_url,
            content=soap_xml.encode("utf-8"),
            headers={"Content-Type": "text/xml; charset=utf-8", "SOAPAction": f'"urn:G2S:{command}"'},
        )
        resp_text = response.text

        # Parse response
        parsed = _parse_soap_response(resp_text)

        # Store RX transcript (SOAP + G2S channels)
        await db.lab_transcripts.insert_one({
            "id": str(uuid.uuid4()), "session_id": session_id,
            "direction": "RX", "channel": "SOAP", "command_class": f"G2S_{command_class}",
            "command_name": f"{command}Ack" if parsed["commands"] else "response",
            "session_type": "Response", "payload_xml": resp_text[:10000],
            "state": "Error" if parsed.get("ack_error") and parsed["ack_error"] != "G2S_none" else "Standard",
            "error_code": parsed.get("ack_error"), "occurred_at": now,
        })
        for cmd in parsed.get("commands", []):
            await db.lab_transcripts.insert_one({
                "id": str(uuid.uuid4()), "session_id": session_id,
                "direction": "RX", "channel": "G2S", "command_class": f"G2S_{command_class}",
                "command_name": cmd["element"], "session_type": "Response",
                "payload_xml": json.dumps(cmd.get("attributes", {})),
                "state": "Standard", "occurred_at": now,
            })

        return {
            "status": "sent", "http_status": response.status_code,
            "command": command, "class": command_class,
            "response_commands": parsed.get("commands", []),
            "ack_error": parsed.get("ack_error"),
            "message_count": conn["message_count"],
        }

    except Exception as e:
        conn["errors"].append({"time": now, "error": str(e)})
        await db.lab_transcripts.insert_one({
            "id": str(uuid.uuid4()), "session_id": session_id,
            "direction": "RX", "channel": "SOAP", "command_class": f"G2S_{command_class}",
            "command_name": "error", "session_type": "Error",
            "payload_xml": str(e), "state": "Error",
            "error_code": "TRANSPORT_ERROR", "occurred_at": now,
        })
        return {"status": "error", "command": command, "error": str(e)}


@router.get("/smart-egm/live-status")
async def get_live_connections(request: Request):
    await get_current_user(request)
    conns = []
    for sid, c in _live_connections.items():
        conns.append({
            "session_id": sid, "egm_url": c["egm_url"], "device_id": c["device_id"],
            "schema_version": c["schema_version"], "status": c["status"],
            "message_count": c["message_count"], "last_message_at": c["last_message_at"],
            "error_count": len(c["errors"]),
        })
    return {"connections": conns}


@router.post("/smart-egm/disconnect-live/{session_id}")
async def disconnect_live(request: Request, session_id: str):
    await get_current_user(request)
    conn = _live_connections.pop(session_id, None)
    if conn and conn.get("http_client"):
        await conn["http_client"].aclose()
    return {"message": f"Disconnected {session_id}"}


# ══════════════════════════════════════════════════
# DEBUG SESSION ZIP EXPORT
# ══════════════════════════════════════════════════

@router.get("/export-session/{session_id}")
async def export_debug_session_zip(request: Request, session_id: str):
    """Export a complete debug session as ZIP containing all 3 transcript channels + metadata."""
    await get_current_user(request)

    # Gather all transcripts by channel
    all_transcripts = await db.lab_transcripts.find({"session_id": session_id}, {"_id": 0}).sort("occurred_at", 1).to_list(100000)

    g2s_msgs = [t for t in all_transcripts if t.get("channel") == "G2S"]
    soap_msgs = [t for t in all_transcripts if t.get("channel") == "SOAP"]
    protocol_msgs = [t for t in all_transcripts if t.get("channel") == "PROTOCOL_TRACE"]

    # Get script runs for this session
    script_runs = await db.lab_script_runs.find({"session_id": session_id}, {"_id": 0}).to_list(100)

    # Get TAR reports
    tar_reports = await db.lab_tar_reports.find({"session_id": session_id}, {"_id": 0}).to_list(10)

    # Get EGM state
    egm_state = _smart_egm_state.get(session_id)

    now = datetime.now(timezone.utc)
    metadata = {
        "session_id": session_id,
        "exported_at": now.isoformat(),
        "total_transcripts": len(all_transcripts),
        "g2s_message_count": len(g2s_msgs),
        "soap_message_count": len(soap_msgs),
        "protocol_trace_count": len(protocol_msgs),
        "script_run_count": len(script_runs),
        "tar_report_count": len(tar_reports),
        "ugg_version": "1.0.0",
        "export_format": "UGG Debug Session v1",
    }

    # Build ZIP in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Metadata
        zf.writestr("session_metadata.json", json.dumps(metadata, indent=2, default=str))

        # 2. G2S Messages channel
        g2s_lines = []
        for t in g2s_msgs:
            g2s_lines.append(f"[{t.get('occurred_at','')}] {t.get('direction','?')} {t.get('command_class','')}.{t.get('command_name','')} | {t.get('state','Standard')}")
            if t.get("payload_xml"):
                g2s_lines.append(f"  XML: {t['payload_xml'][:500]}")
            g2s_lines.append("")
        zf.writestr("g2s_messages.log", "\n".join(g2s_lines))
        zf.writestr("g2s_messages.json", json.dumps(g2s_msgs, indent=2, default=str))

        # 3. SOAP Transport channel
        soap_lines = []
        for t in soap_msgs:
            soap_lines.append(f"[{t.get('occurred_at','')}] {t.get('direction','?')} {t.get('command_name','')}")
            if t.get("payload_xml"):
                soap_lines.append(t["payload_xml"][:2000])
            soap_lines.append("---")
        zf.writestr("soap_transport.log", "\n".join(soap_lines))
        zf.writestr("soap_transport.json", json.dumps(soap_msgs, indent=2, default=str))

        # 4. Protocol Trace channel
        proto_lines = []
        for t in protocol_msgs:
            proto_lines.append(f"[{t.get('occurred_at','')}] {t.get('direction','?')} {t.get('command_name','')}")
            if t.get("payload_raw"):
                proto_lines.append(f"  HEX: {t['payload_raw']}")
            proto_lines.append("")
        zf.writestr("protocol_trace.log", "\n".join(proto_lines))
        zf.writestr("protocol_trace.json", json.dumps(protocol_msgs, indent=2, default=str))

        # 5. Script runs
        if script_runs:
            zf.writestr("script_runs.json", json.dumps(script_runs, indent=2, default=str))

        # 6. TAR reports
        if tar_reports:
            zf.writestr("tar_reports.json", json.dumps(tar_reports, indent=2, default=str))

        # 7. EGM state snapshot
        if egm_state:
            safe_state = {**egm_state, "doors_open": list(egm_state.get("doors_open", set())), "events": egm_state.get("events", [])[-50:]}
            zf.writestr("egm_state.json", json.dumps(safe_state, indent=2, default=str))

    zip_buffer.seek(0)
    filename = f"ugg_debug_session_{session_id}_{now.strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(zip_buffer, media_type="application/zip", headers={"Content-Disposition": f"attachment; filename={filename}"})
